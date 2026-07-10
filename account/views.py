from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.utils import timezone
import json
from decimal import Decimal
from .forms import RegistrationForm, LoginForm, CompanySetupForm, CompanyProfileForm, BusinessProfileForm, ItemForm, CustomerForm, InvoiceForm, CustomerDetailsForm, VendorForm, BillForm, VendorDetailsForm, AccountForm, IncomeTransactionForm, ExpenseTransactionForm, TransferForm, ReconciliationForm, ReportForm, ScheduledReportForm, ReportEditForm, DashboardWidgetForm, DashboardForm
from .models import Company, Category, Tax, IncomeAccount, ExpenseAccount, Item, Customer, InvoiceCategory, Invoice, InvoiceItem, InvoiceAttachment, CustomerContact, Vendor, Bill, BillItem, VendorContact, Account, IncomeTransaction, ExpenseTransaction, Transfer, Reconciliation, ReconciliationTransaction, Report, PinnedReport, ScheduledReport, App, AppInstallation, DashboardWidget, Dashboard, Notification


def is_google_configured(request):
    try:
        from allauth.socialaccount.adapter import get_adapter
        get_adapter(request).get_app(request, 'google')
        return True
    except Exception:
        return False


def landing(request):
    if request.user.is_authenticated:
        try:
            company = request.user.company
            if company.onboarding_step == 1:
                return redirect('company_profile')
            elif company.onboarding_step == 2:
                return redirect('business_profile')
            elif company.onboarding_step == 3:
                return redirect('dashboard')
            return redirect('company_setup')
        except Company.DoesNotExist:
            return redirect('company_setup')
    return render(request, 'authentication/landing.html')


def login_view(request):
    if request.user.is_authenticated:
        return redirect('landing')
        
    google_configured = is_google_configured(request)
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username_or_email = form.cleaned_data['username']
            password = form.cleaned_data['password']
            remember_me = form.cleaned_data['remember_me']
            
            user = authenticate(request, username=username_or_email, password=password)
            if user is not None:
                login(request, user)
                if not remember_me:
                    request.session.set_expiry(0)
                messages.success(request, f"Welcome back, {user.first_name or user.username}!")
                return redirect('landing')
            else:
                messages.error(request, "Invalid username/email or password.")
    else:
        form = LoginForm()
    return render(request, 'authentication/login.html', {'form': form, 'google_configured': google_configured})


def logout_view(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('landing')


def register_view(request):
    if request.user.is_authenticated:
        return redirect('landing')
        
    google_configured = is_google_configured(request)
    
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.save()
            
            login(request, user, backend='account.backends.EmailOrUsernameModelBackend')
            messages.success(request, "Registration successful!")
            return redirect('company_setup')
    else:
        form = RegistrationForm()
    return render(request, 'authentication/register.html', {'form': form, 'google_configured': google_configured})


@login_required
def company_setup_view(request):
    try:
        company = request.user.company
        if company.onboarding_step == 1:
            return redirect('company_profile')
        elif company.onboarding_step == 2:
            return redirect('business_profile')
        elif company.onboarding_step == 3:
            return redirect('onboarding_complete')
    except Company.DoesNotExist:
        company = None
        
    if request.method == 'POST':
        form = CompanySetupForm(request.POST)
        if form.is_valid():
            company = form.save(commit=False)
            company.user = request.user
            company.onboarding_step = 1
            company.save()
            return redirect('loading')
    else:
        form = CompanySetupForm(initial={'language': 'en', 'currency': 'USD'})
        
    return render(request, 'authentication/company_setup.html', {'form': form})


@login_required
def loading_view(request):
    try:
        company = request.user.company
        if company.onboarding_step != 1:
            return redirect('landing')
    except Company.DoesNotExist:
        return redirect('company_setup')
    return render(request, 'authentication/loading.html')


@login_required
def company_profile_wizard(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
        
    if request.method == 'POST':
        if 'skip' in request.POST or 'skip' in request.GET:
            company.onboarding_step = 2
            company.save()
            return redirect('business_profile')
            
        form = CompanyProfileForm(request.POST, request.FILES, instance=company)
        if form.is_valid():
            company = form.save(commit=False)
            company.onboarding_step = 2
            company.save()
            return redirect('business_profile')
    else:
        form = CompanyProfileForm(instance=company)
        
    return render(request, 'authentication/company_profile.html', {'form': form, 'company': company})


@login_required
def business_profile_wizard(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
        
    if request.method == 'POST':
        if 'previous' in request.POST:
            company.onboarding_step = 1
            company.save()
            return redirect('company_profile')
            
        form = BusinessProfileForm(request.POST, instance=company)
        if form.is_valid():
            company = form.save(commit=False)
            company.onboarding_step = 3
            company.save()
            return redirect('onboarding_complete')
    else:
        form = BusinessProfileForm(instance=company)
        
    return render(request, 'authentication/business_profile.html', {'form': form, 'company': company})


@login_required
def onboarding_complete_view(request):
    try:
        company = request.user.company
        if company.onboarding_step < 3:
            return redirect('landing')
    except Company.DoesNotExist:
        return redirect('company_setup')
    return render(request, 'authentication/onboarding_complete.html', {'company': company})


@login_required
def dashboard(request):
    try:
        company = request.user.company
        if company.onboarding_step < 3:
            return redirect('landing')
    except Company.DoesNotExist:
        return redirect('company_setup')

    dashboards = Dashboard.objects.filter(users=request.user, is_active=True).order_by('-created_at')

    if not dashboards.exists():
        default_db, created = Dashboard.objects.get_or_create(
            name="Default Dashboard",
            created_by=request.user,
            defaults={'is_active': True}
        )
        default_db.users.add(request.user)
        dashboards = Dashboard.objects.filter(users=request.user, is_active=True).order_by('-created_at')

    selected_db_id = request.GET.get('dashboard_id')
    if selected_db_id:
        active_db = get_object_or_404(Dashboard, id=selected_db_id, users=request.user, is_active=True)
    else:
        active_db = dashboards.first()

    if not DashboardWidget.objects.filter(user=request.user, dashboard=active_db).exists():
        defaults = [
            ('Receivables', 'Card', 50, 0),
            ('Payables', 'Card', 50, 1),
            ('Cash Flow', 'Chart', 100, 2),
            ('Profit & Loss', 'Chart', 50, 3),
            ('Expenses By Category', 'Chart', 50, 4),
            ('Account Balance', 'List', 50, 5),
            ('Connect Bank Accounts', 'Summary', 50, 6),
            ('Recent Activity', 'List', 50, 7),
        ]
        for name, wtype, width, order in defaults:
            DashboardWidget.objects.get_or_create(
                user=request.user, dashboard=active_db, widget_name=name,
                defaults={'widget_type': wtype, 'width': width, 'sort_order': order}
            )

    # Make sure 'Recent Activity' widget is present in the dashboard widgets list
    DashboardWidget.objects.get_or_create(
        user=request.user, dashboard=active_db, widget_name='Recent Activity',
        defaults={'widget_type': 'List', 'width': 50, 'sort_order': 7}
    )

    widgets = DashboardWidget.objects.filter(user=request.user, dashboard=active_db, is_active=True).order_by('sort_order')

    # --- Live Dashboard ORM Calculations ---
    from django.db.models import Sum
    from django.db.models.functions import ExtractMonth
    import datetime

    today = timezone.now().date()
    current_year = today.year

    # 1. Receivables (Invoices)
    invoices = Invoice.objects.filter(company=company)
    total_sales_amount = invoices.aggregate(Sum('total'))['total__sum'] or 0.00
    overdue_receivables = invoices.filter(due_date__lt=today).aggregate(Sum('total'))['total__sum'] or 0.00
    open_receivables = invoices.filter(due_date__gte=today).aggregate(Sum('total'))['total__sum'] or 0.00
    total_receivables = float(total_sales_amount)
    
    receivables_progress = 0
    if total_receivables > 0:
        receivables_progress = int((float(open_receivables) / total_receivables) * 100)

    # 2. Payables (Bills)
    bills = Bill.objects.filter(company=company)
    total_payables_amount = bills.aggregate(Sum('total'))['total__sum'] or 0.00
    overdue_payables = bills.filter(due_date__lt=today).aggregate(Sum('total'))['total__sum'] or 0.00
    open_payables = bills.filter(due_date__gte=today).aggregate(Sum('total'))['total__sum'] or 0.00
    total_payables = float(total_payables_amount)
    
    payables_progress = 0
    if total_payables > 0:
        payables_progress = int((float(open_payables) / total_payables) * 100)

    # 3. Cash Flow (Transactions)
    income_txs = IncomeTransaction.objects.filter(company=company)
    expense_txs = ExpenseTransaction.objects.filter(company=company)
    
    total_incoming = income_txs.aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    total_outgoing = expense_txs.aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    net_profit = total_incoming - total_outgoing

    # Monthly Cash Flow (Jan-Dec) for ApexCharts
    monthly_income = income_txs.filter(date__year=current_year)\
        .annotate(month=ExtractMonth('date'))\
        .values('month')\
        .annotate(total=Sum('amount'))\
        .order_by('month')
        
    monthly_expense = expense_txs.filter(date__year=current_year)\
        .annotate(month=ExtractMonth('date'))\
        .values('month')\
        .annotate(total=Sum('amount'))\
        .order_by('month')

    income_data = [0.0] * 12
    expense_data = [0.0] * 12
    net_flow_data = [0.0] * 12

    for item in monthly_income:
        m = item['month']
        if 1 <= m <= 12:
            income_data[m-1] = float(item['total'])

    for item in monthly_expense:
        m = item['month']
        if 1 <= m <= 12:
            expense_data[m-1] = float(item['total'])

    for i in range(12):
        net_flow_data[i] = income_data[i] - expense_data[i]

    # 4. Profit & Loss (Last 6 Months)
    pl_labels = []
    pl_series_data = []
    for i in range(5, -1, -1):
        first_day_of_month = (today.replace(day=1) - datetime.timedelta(days=i*30)).replace(day=1)
        m_num = first_day_of_month.month
        y_num = first_day_of_month.year
        pl_labels.append(first_day_of_month.strftime('%b'))
        
        inc = income_txs.filter(date__year=y_num, date__month=m_num).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        exp = expense_txs.filter(date__year=y_num, date__month=m_num).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        pl_series_data.append(float(inc - exp))

    # 5. Expenses by Category (Top 5)
    category_expenses = expense_txs.values('category__name')\
        .annotate(total=Sum('amount'))\
        .order_by('-total')[:5]

    exp_labels = [item['category__name'] or 'General/Uncategorized' for item in category_expenses]
    exp_series = [float(item['total']) for item in category_expenses]
    if not exp_series:
        exp_labels = ['No Expenses']
        exp_series = [0.0]

    # 6. Account Balances (Real-time recalculation)
    accounts = Account.objects.filter(company=company)
    dashboard_accounts = []
    for acc in accounts:
        inc_sum = income_txs.filter(account=acc).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        exp_sum = expense_txs.filter(account=acc).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        from_transfers = Transfer.objects.filter(company=company, from_account=acc).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        to_transfers = Transfer.objects.filter(company=company, to_account=acc).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        
        balance = float(acc.opening_balance) + float(inc_sum) - float(exp_sum) - float(from_transfers) + float(to_transfers)
        dashboard_accounts.append({
            'name': acc.name,
            'balance': balance
        })

    # 7. Recent Activities (Merged latest 10 from Notification table)
    activities = []
    for notif in Notification.objects.filter(recipient=request.user, company__name=company.name).order_by('-timestamp')[:10]:
        activities.append({
            'time': notif.timestamp,
            'text': notif.message
        })
    recent_activities = activities

    return render(request, 'dashboard/dashboard.html', {
        'company': company,
        'widgets': widgets,
        'dashboards': dashboards,
        'active_db': active_db,
        
        # Receivables
        'total_receivables': total_receivables,
        'receivables_progress': receivables_progress,
        'open_receivables': open_receivables,
        'overdue_receivables': overdue_receivables,
        
        # Payables
        'total_payables': total_payables,
        'payables_progress': payables_progress,
        'open_payables': open_payables,
        'overdue_payables': overdue_payables,
        
        # Cash Flow
        'total_incoming': total_incoming,
        'total_outgoing': total_outgoing,
        'net_profit': net_profit,
        'net_flow_data': net_flow_data,
        
        # Profit & Loss
        'pl_labels': pl_labels,
        'pl_series_data': pl_series_data,
        
        # Expenses By Category
        'exp_labels': exp_labels,
        'exp_series': exp_series,
        
        # Account Balances & Activities
        'dashboard_accounts': dashboard_accounts,
        'recent_activities': recent_activities,
    })



@login_required
def items_list_view(request):
    try:
        company = request.user.company
        if company.onboarding_step < 3:
            return redirect('landing')
    except Company.DoesNotExist:
        return redirect('company_setup')
    from account.models import Item
    items = Item.objects.filter(company__name=company.name)
    return render(request, 'items/index.html', {'company': company, 'items': items})


@login_required
def item_create_view(request):
    try:
        company = request.user.company
        if company.onboarding_step < 3:
            return redirect('landing')
    except Company.DoesNotExist:
        return redirect('company_setup')
        
    # Programmatically populate defaults if they are empty
    if not Category.objects.exists():
        Category.objects.create(name="Products", description="Physical goods")
        Category.objects.create(name="Services", description="Service-based items")
        
    if not Tax.objects.exists():
        Tax.objects.create(tax_name="VAT 5%", tax_percentage=5.00)
        Tax.objects.create(tax_name="GST 18%", tax_percentage=18.00)
        
    if not IncomeAccount.objects.exists():
        IncomeAccount.objects.create(account_name="Sales", account_code="400")
        IncomeAccount.objects.create(account_name="Interest Income", account_code="410")
        
    if not ExpenseAccount.objects.exists():
        ExpenseAccount.objects.create(account_name="General Expenses", account_code="628")
        ExpenseAccount.objects.create(account_name="Cost of Goods Sold", account_code="600")

    if request.method == 'POST':
        form = ItemForm(request.POST, company=company)
        if form.is_valid():
            item = form.save(commit=False)
            item.company = company
            item.created_by = request.user
            item.save()
            messages.success(request, "Item Created Successfully")
            return redirect('items_list')
    else:
        form = ItemForm(company=company, initial={'type': 'product', 'sale_enabled': True, 'purchase_enabled': True})
        
    return render(request, 'items/new_item.html', {'form': form, 'company': company})


@login_required
def item_edit_view(request, id):
    try:
        company = request.user.company
        if company.onboarding_step < 3:
            return redirect('landing')
    except Company.DoesNotExist:
        return redirect('company_setup')
        
    item = get_object_or_404(Item, id=id, company__name=company.name)
    
    if request.method == 'POST':
        form = ItemForm(request.POST, instance=item, company=company)
        if form.is_valid():
            form.save()
            messages.success(request, "Item Updated Successfully")
            return redirect('items_list')
    else:
        form = ItemForm(instance=item, company=company)
        
    return render(request, 'items/new_item.html', {'form': form, 'company': company})


@login_required
def item_detail_view(request, id):
    try:
        company = request.user.company
        if company.onboarding_step < 3:
            return redirect('landing')
    except Company.DoesNotExist:
        return redirect('company_setup')
        
    item = get_object_or_404(Item, id=id, company__name=company.name)
    return render(request, 'items/item_detail.html', {'item': item, 'company': company})


@login_required
def item_import_view(request):
    try:
        company = request.user.company
        if company.onboarding_step < 3:
            return redirect('landing')
    except Company.DoesNotExist:
        return redirect('company_setup')
        
    if request.method == 'POST':
        import_file = request.FILES.get('import_file')
        if not import_file:
            messages.error(request, "Please select an Excel file to import.")
            return redirect('item_import')
            
        import openpyxl
        try:
            wb = openpyxl.load_workbook(import_file, read_only=True)
            sheet = wb.active
            
            imported = 0
            skipped = 0
            failed = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                    
                name = row[0]
                item_type = str(row[1]).lower() if row[1] else 'product'
                sale_price = row[2]
                purchase_price = row[3]
                description = row[4] if len(row) > 4 else ""
                
                if item_type not in ['product', 'service']:
                    item_type = 'product'
                    
                if not name:
                    failed += 1
                    continue
                    
                if Item.objects.filter(company=company, name__iexact=name).exists():
                    skipped += 1
                    continue
                    
                try:
                    Item.objects.create(
                        company=company,
                        name=name,
                        type=item_type,
                        sale_price=sale_price if sale_price is not None else 0.00,
                        purchase_price=purchase_price if purchase_price is not None else 0.00,
                        description=description,
                        created_by=request.user
                    )
                    imported += 1
                except Exception:
                    failed += 1
            
            messages.success(request, f"Import Summary - Imported: {imported}, Skipped: {skipped}, Failed: {failed}")
            return redirect('items_list')
            
        except Exception as e:
            messages.error(request, f"Failed to read Excel file: {str(e)}")
            return redirect('item_import')
            
    return render(request, 'items/import_items.html', {'company': company})


@login_required
def inventory_view(request):
    try:
        company = request.user.company
        if company.onboarding_step < 3:
            return redirect('landing')
    except Company.DoesNotExist:
        return redirect('company_setup')
        
    items = Item.objects.filter(company__name=company.name)
    return render(request, 'items/inventory.html', {'company': company, 'items': items})


from django.http import JsonResponse
from django.utils import timezone
import json

@login_required
def invoices_list_view(request):
    try:
        company = request.user.company
        if company.onboarding_step < 3:
            return redirect('landing')
    except Company.DoesNotExist:
        return redirect('company_setup')
    invoices = Invoice.objects.filter(company=company)
    return render(request, 'invoices/invoice_list.html', {'company': company, 'invoices': invoices})


@login_required
def invoice_create_view(request):
    try:
        company = request.user.company
        if company.onboarding_step < 3:
            return redirect('landing')
    except Company.DoesNotExist:
        return redirect('company_setup')

    if not InvoiceCategory.objects.exists():
        InvoiceCategory.objects.create(name="Sales", status="Active")
        InvoiceCategory.objects.create(name="Consulting", status="Active")

    last_invoice = Invoice.objects.filter(company=company).order_by('id').last()
    if last_invoice and last_invoice.invoice_number.startswith('INV-'):
        try:
            num = int(last_invoice.invoice_number.split('-')[1])
            next_num = num + 1
        except Exception:
            next_num = 1
    else:
        next_num = 1
    next_invoice_number = f"INV-{next_num:06d}"

    if request.method == 'POST':
        form = InvoiceForm(request.POST, company=company)
        if form.is_valid():
            from django.db import transaction
            try:
                with transaction.atomic():
                    invoice = form.save(commit=False)
                    invoice.company = company
                    invoice.created_by = request.user
                    
                    invoice.subtotal = request.POST.get('subtotal', 0.00)
                    invoice.discount = request.POST.get('discount', 0.00)
                    invoice.tax_amount = request.POST.get('tax_amount', 0.00)
                    invoice.total = request.POST.get('total', 0.00)
                    invoice.save()

                    items_json = request.POST.get('items_data')
                    if items_json:
                        items_data = json.loads(items_json)
                        for item_data in items_data:
                            item_id = item_data.get('item_id')
                            item_obj = Item.objects.filter(id=item_id).first() if item_id else None
                            InvoiceItem.objects.create(
                                invoice=invoice,
                                item=item_obj,
                                name=item_data.get('name', ''),
                                description=item_data.get('description', ''),
                                quantity=item_data.get('quantity', 1),
                                price=item_data.get('price', 0.00),
                                amount=item_data.get('amount', 0.00)
                            )

                    files = request.FILES.getlist('attachment_files')
                    for f in files:
                        InvoiceAttachment.objects.create(invoice=invoice, file=f)

                messages.success(request, "Invoice Created Successfully")
                return redirect('invoices_list')
            except Exception as e:
                messages.error(request, f"Error saving invoice: {str(e)}")
    else:
        form = InvoiceForm(company=company, initial={
            'invoice_number': next_invoice_number,
            'invoice_date': timezone.now().strftime('%Y-%m-%d'),
            'due_date': timezone.now().strftime('%Y-%m-%d')
        })
    
    customers = Customer.objects.filter(company__name=company.name)
    items = Item.objects.filter(company__name=company.name)
    categories = InvoiceCategory.objects.filter(status='Active')
    customer_form = CustomerForm()

    return render(request, 'invoices/invoice_form.html', {
        'form': form,
        'company': company,
        'customers': customers,
        'items': items,
        'categories': categories,
        'customer_form': customer_form
    })


@login_required
def ajax_customer_create(request):
    if request.method == 'POST':
        try:
            company = request.user.company
        except Company.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'No company found'}, status=400)
            
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save(commit=False)
            customer.company = company
            customer.save()
            return JsonResponse({'success': True, 'id': customer.id, 'name': customer.name})
        else:
            return JsonResponse({'success': False, 'errors': form.errors.as_json()}, status=400)
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)


@login_required
def invoice_detail_view(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
    invoice = Invoice.objects.filter(id=id, company=company).first()
    if not invoice:
        messages.error(request, "Invoice not found.")
        return redirect('invoices_list')
    return render(request, 'invoices/invoice_detail.html', {'company': company, 'invoice': invoice})


@login_required
def invoice_edit_view(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
    invoice = Invoice.objects.filter(id=id, company=company).first()
    if not invoice:
        messages.error(request, "Invoice not found.")
        return redirect('invoices_list')

    if request.method == 'POST':
        form = InvoiceForm(request.POST, instance=invoice, company=company)
        if form.is_valid():
            from django.db import transaction
            try:
                with transaction.atomic():
                    invoice = form.save(commit=False)
                    invoice.subtotal = request.POST.get('subtotal', 0.00)
                    invoice.discount = request.POST.get('discount', 0.00)
                    invoice.tax_amount = request.POST.get('tax_amount', 0.00)
                    invoice.total = request.POST.get('total', 0.00)
                    invoice.save()

                    invoice.items.all().delete()
                    items_json = request.POST.get('items_data')
                    if items_json:
                        items_data = json.loads(items_json)
                        for item_data in items_data:
                            item_id = item_data.get('item_id')
                            item_obj = Item.objects.filter(id=item_id).first() if item_id else None
                            InvoiceItem.objects.create(
                                invoice=invoice,
                                item=item_obj,
                                name=item_data.get('name', ''),
                                description=item_data.get('description', ''),
                                quantity=item_data.get('quantity', 1),
                                price=item_data.get('price', 0.00),
                                amount=item_data.get('amount', 0.00)
                            )

                    files = request.FILES.getlist('attachment_files')
                    for f in files:
                        InvoiceAttachment.objects.create(invoice=invoice, file=f)

                messages.success(request, "Invoice Updated Successfully")
                return redirect('invoice_detail', id=invoice.id)
            except Exception as e:
                messages.error(request, f"Error updating invoice: {str(e)}")
    else:
        form = InvoiceForm(instance=invoice, company=company)
        
    customers = Customer.objects.filter(company__name=company.name)
    items = Item.objects.filter(company__name=company.name)
    categories = InvoiceCategory.objects.filter(status='Active')
    customer_form = CustomerForm()

    return render(request, 'invoices/invoice_form.html', {
        'form': form,
        'company': company,
        'invoice': invoice,
        'customers': customers,
        'items': items,
        'categories': categories,
        'customer_form': customer_form
    })


@login_required
def invoice_delete_view(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
    invoice = Invoice.objects.filter(id=id, company=company).first()
    if invoice:
        invoice.delete()
        messages.success(request, "Invoice Deleted Successfully")
    else:
        messages.error(request, "Invoice not found.")
    return redirect('invoices_list')


@login_required
def invoice_import_view(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
        
    if request.method == 'POST':
        import_file = request.FILES.get('import_file')
        if not import_file:
            messages.error(request, "Please select an Excel file to import.")
            return redirect('invoice_import')
            
        import openpyxl
        try:
            wb = openpyxl.load_workbook(import_file, read_only=True)
            sheet = wb.active
            
            imported = 0
            skipped = 0
            failed = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                    
                customer_name = row[0]
                invoice_number = row[1]
                invoice_date_val = row[2]
                due_date_val = row[3]
                total_val = row[4]
                notes_val = row[5] if len(row) > 5 else ""
                
                if not customer_name or not invoice_number:
                    failed += 1
                    continue
                    
                if Invoice.objects.filter(company=company, invoice_number=invoice_number).exists():
                    skipped += 1
                    continue
                    
                customer, created = Customer.objects.get_or_create(company=company, name=customer_name)
                
                try:
                    Invoice.objects.create(
                        company=company,
                        customer=customer,
                        invoice_number=invoice_number,
                        invoice_date=invoice_date_val if invoice_date_val else timezone.now().date(),
                        due_date=due_date_val if due_date_val else timezone.now().date(),
                        total=total_val if total_val else 0.00,
                        subtotal=total_val if total_val else 0.00,
                        notes=notes_val,
                        created_by=request.user
                    )
                    imported += 1
                except Exception:
                    failed += 1
                    
            messages.success(request, f"Import Summary - Imported: {imported}, Skipped: {skipped}, Failed: {failed}")
            return redirect('invoices_list')
        except Exception as e:
            messages.error(request, f"Failed to read Excel file: {str(e)}")
            return redirect('invoice_import')
            
    return render(request, 'invoices/invoice_import.html', {'company': company})


@login_required
def customer_home_view(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
    from account.models import Customer
    customers = Customer.objects.filter(company=company)
    if customers.exists():
        return redirect('customers_list')
    return render(request, 'customers/customer_home.html', {'company': company})


@login_required
def customer_create_view(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
        
    if request.method == 'POST':
        form = CustomerDetailsForm(request.POST, company=company)
        if form.is_valid():
            customer = form.save(commit=False)
            customer.company = company
            customer.save()
            
            contact_names = request.POST.getlist('contact_name[]')
            contact_emails = request.POST.getlist('contact_email[]')
            contact_phones = request.POST.getlist('contact_phone[]')
            
            for i in range(len(contact_names)):
                name = contact_names[i]
                if name:
                    CustomerContact.objects.create(
                        customer=customer,
                        name=name,
                        email=contact_emails[i] if i < len(contact_emails) else "",
                        phone=contact_phones[i] if i < len(contact_phones) else ""
                    )
            
            messages.success(request, "Customer created successfully.")
            return redirect('customers_list')
    else:
        form = CustomerDetailsForm(company=company)
        
    return render(request, 'customers/customer_form.html', {'form': form, 'company': company})


@login_required
def customers_list_view(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
        
    query = request.GET.get('q', '')
    sort_by = request.GET.get('sort', '-created_at')
    
    customers = Customer.objects.filter(company=company)
    if query:
        customers = customers.filter(name__icontains=query)
        
    allowed_sorts = ['name', '-name', 'email', '-email', 'phone', '-phone', 'reference', '-reference', 'currency', '-currency', 'created_at', '-created_at']
    if sort_by in allowed_sorts:
        customers = customers.order_by(sort_by)
        
    from django.core.paginator import Paginator
    paginator = Paginator(customers, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'customers/customer_list.html', {
        'company': company,
        'page_obj': page_obj,
        'query': query,
        'sort_by': sort_by
    })


@login_required
def customer_edit_view(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
        
    customer = Customer.objects.filter(id=id, company=company).first()
    if not customer:
        messages.error(request, "Customer not found.")
        return redirect('customers_list')
        
    if request.method == 'POST':
        form = CustomerDetailsForm(request.POST, instance=customer, company=company)
        if form.is_valid():
            form.save()
            
            customer.contacts.all().delete()
            contact_names = request.POST.getlist('contact_name[]')
            contact_emails = request.POST.getlist('contact_email[]')
            contact_phones = request.POST.getlist('contact_phone[]')
            
            for i in range(len(contact_names)):
                name = contact_names[i]
                if name:
                    CustomerContact.objects.create(
                        customer=customer,
                        name=name,
                        email=contact_emails[i] if i < len(contact_emails) else "",
                        phone=contact_phones[i] if i < len(contact_phones) else ""
                    )
            
            messages.success(request, "Customer updated successfully.")
            return redirect('customers_list')
    else:
        form = CustomerDetailsForm(instance=customer, company=company)
        
    return render(request, 'customers/customer_form.html', {
        'form': form,
        'company': company,
        'customer': customer,
        'contacts': customer.contacts.all()
    })


@login_required
def customer_delete_view(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
        
    customer = Customer.objects.filter(id=id, company=company).first()
    if customer:
        customer.delete()
        messages.success(request, "Customer deleted successfully.")
    else:
        messages.error(request, "Customer not found.")
    return redirect('customers_list')


@login_required
def customer_import_view(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
        
    if request.method == 'POST':
        import_file = request.FILES.get('import_file')
        if not import_file:
            messages.error(request, "Please select an Excel file to import.")
            return redirect('customer_import')
            
        import openpyxl
        try:
            wb = openpyxl.load_workbook(import_file, read_only=True)
            sheet = wb.active
            
            imported = 0
            skipped = 0
            failed = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                    
                name = row[0]
                email = row[1]
                phone = row[2]
                currency = row[3] if len(row) > 3 and row[3] else 'INR'
                
                if not name:
                    failed += 1
                    continue
                    
                if Customer.objects.filter(company=company, name__iexact=name).exists():
                    skipped += 1
                    continue
                    
                try:
                    Customer.objects.create(
                        company=company,
                        name=name,
                        email=email,
                        phone=phone,
                        currency=currency
                    )
                    imported += 1
                except Exception:
                    failed += 1
            
            messages.success(request, f"Import Summary - Imported: {imported}, Skipped: {skipped}, Failed: {failed}")
            return redirect('customers_list')
            
        except Exception as e:
            messages.error(request, f"Failed to read Excel file: {str(e)}")
            return redirect('customer_import')
            
    return render(request, 'customers/customer_import.html', {'company': company})


from django.views import View
from django.utils.decorators import method_decorator

@method_decorator(login_required, name='dispatch')
class BillsLandingView(View):
    def get(self, request):
        try:
            company = request.user.company
        except Company.DoesNotExist:
            return redirect('company_setup')
        from account.models import Bill
        if Bill.objects.filter(company=company).exists():
            return redirect('bill_list')
        return render(request, 'purchases/bills/landing.html', {'company': company})


@method_decorator(login_required, name='dispatch')
class NewBillView(View):
    def get(self, request):
        try:
            company = request.user.company
        except Company.DoesNotExist:
            return redirect('company_setup')

        last_bill = Bill.objects.filter(company=company).order_by('id').last()
        if last_bill and last_bill.bill_number.startswith('BILL-'):
            try:
                num = int(last_bill.bill_number.split('-')[1])
                next_num = num + 1
            except Exception:
                next_num = 1
        else:
            next_num = 1
        next_bill_number = f"BILL-{next_num:05d}"

        form = BillForm(company=company, initial={
            'bill_number': next_bill_number,
            'bill_date': timezone.now().strftime('%Y-%m-%d'),
            'due_date': timezone.now().strftime('%Y-%m-%d')
        })
        vendors = Vendor.objects.all()
        items = Item.objects.filter(company=company)
        vendor_form = VendorForm()

        return render(request, 'purchases/bills/new_bill.html', {
            'form': form,
            'company': company,
            'vendors': vendors,
            'items': items,
            'vendor_form': vendor_form
        })


@method_decorator(login_required, name='dispatch')
class SaveBillView(View):
    def post(self, request):
        try:
            company = request.user.company
        except Company.DoesNotExist:
            return redirect('company_setup')

        form = BillForm(request.POST, request.FILES, company=company)
        if form.is_valid():
            from django.db import transaction
            try:
                with transaction.atomic():
                    bill = form.save(commit=False)
                    bill.company = company
                    bill.created_by = request.user
                    
                    bill.subtotal = request.POST.get('subtotal', 0.00)
                    bill.discount = request.POST.get('discount', 0.00)
                    bill.tax = request.POST.get('tax', 0.00)
                    bill.total = request.POST.get('total', 0.00)
                    bill.save()

                    items_json = request.POST.get('items_data')
                    if items_json:
                        items_data = json.loads(items_json)
                        for item_data in items_data:
                            BillItem.objects.create(
                                bill=bill,
                                item_name=item_data.get('name', ''),
                                description=item_data.get('description', ''),
                                quantity=item_data.get('quantity', 1),
                                price=item_data.get('price', 0.00),
                                amount=item_data.get('amount', 0.00)
                            )
                messages.success(request, "Bill created successfully.")
                return redirect('bill_list')
            except Exception as e:
                messages.error(request, f"Error saving bill: {str(e)}")
                items = Item.objects.filter(company=company)
                vendor_form = VendorDetailsForm(company=company)
                return render(request, 'purchases/bills/new_bill.html', {
                    'form': form,
                    'company': company,
                    'items': items,
                    'vendor_form': vendor_form
                })
        else:
            messages.error(request, "Please correct the form errors.")
            items = Item.objects.filter(company=company)
            vendor_form = VendorDetailsForm(company=company)
            return render(request, 'purchases/bills/new_bill.html', {
                'form': form,
                'company': company,
                'items': items,
                'vendor_form': vendor_form
            })


@method_decorator(login_required, name='dispatch')
class ImportBillView(View):
    def get(self, request):
        try:
            company = request.user.company
        except Company.DoesNotExist:
            return redirect('company_setup')
        return render(request, 'purchases/bills/import_bill.html', {'company': company})


@method_decorator(login_required, name='dispatch')
class ImportBillExcelView(View):
    def post(self, request):
        try:
            company = request.user.company
        except Company.DoesNotExist:
            return redirect('company_setup')

        import_file = request.FILES.get('import_file')
        if not import_file:
            messages.error(request, "Please select an Excel file to import.")
            return redirect('bill_import')

        import openpyxl
        try:
            wb = openpyxl.load_workbook(import_file, read_only=True)
            sheet = wb.active

            imported = 0
            skipped = 0
            failed = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue

                vendor_name = row[0]
                bill_number = row[1]
                bill_date_val = row[2]
                due_date_val = row[3]
                total_val = row[4]
                notes_val = row[5] if len(row) > 5 else ""

                if not vendor_name or not bill_number:
                    failed += 1
                    continue

                if Bill.objects.filter(company=company, bill_number=bill_number).exists():
                    skipped += 1
                    continue

                vendor, created = Vendor.objects.get_or_create(vendor_name=vendor_name)

                try:
                    Bill.objects.create(
                        company=company,
                        vendor=vendor,
                        bill_number=bill_number,
                        bill_date=bill_date_val if bill_date_val else timezone.now().date(),
                        due_date=due_date_val if due_date_val else timezone.now().date(),
                        total=total_val if total_val else 0.00,
                        subtotal=total_val if total_val else 0.00,
                        notes=notes_val,
                        created_by=request.user
                    )
                    imported += 1
                except Exception:
                    failed += 1

            messages.success(request, f"{imported} Bills Imported Successfully. {skipped} Duplicate Bills Skipped.")
            return redirect('bill_list')
        except Exception as e:
            messages.error(request, f"Failed to read Excel file: {str(e)}")
            return redirect('bill_import')


@method_decorator(login_required, name='dispatch')
class BillListView(View):
    def get(self, request):
        try:
            company = request.user.company
        except Company.DoesNotExist:
            return redirect('company_setup')
        bills = Bill.objects.filter(company=company)
        return render(request, 'purchases/bills/bill_list.html', {'company': company, 'bills': bills})


@login_required
def vendor_home_view(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
    from account.models import Vendor
    if Vendor.objects.filter(company=company).exists():
        return redirect('vendors_list')
    return render(request, 'vendors/index.html', {'company': company})


@login_required
def vendor_create_view(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    if request.method == 'POST':
        form = VendorDetailsForm(request.POST, request.FILES, company=company)
        if form.is_valid():
            vendor = form.save(commit=False)
            vendor.company = company
            vendor.created_by = request.user
            vendor.save()

            names = request.POST.getlist('contact_name[]')
            emails = request.POST.getlist('contact_email[]')
            phones = request.POST.getlist('contact_phone[]')

            for i in range(len(names)):
                if names[i]:
                    VendorContact.objects.create(
                        vendor=vendor,
                        name=names[i],
                        email=emails[i] if i < len(emails) else "",
                        phone=phones[i] if i < len(phones) else ""
                    )

            messages.success(request, "Vendor created successfully.")
            return redirect('vendors_list')
    else:
        form = VendorDetailsForm(company=company)

    return render(request, 'vendors/create.html', {
        'form': form,
        'company': company
    })


@login_required
def vendors_list_view(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    query = request.GET.get('q', '')
    sort_by = request.GET.get('sort', 'vendor_name')

    vendors = Vendor.objects.filter(company=company)
    if query:
        vendors = vendors.filter(vendor_name__icontains=query)

    if sort_by.endswith('vendor_name') or sort_by == '-vendor_name':
        vendors = vendors.order_by(sort_by)

    paginator = Paginator(vendors, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'vendors/list.html', {
        'company': company,
        'page_obj': page_obj,
        'query': query,
        'sort_by': sort_by
    })


@login_required
def vendor_edit_view(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    vendor = get_object_or_404(Vendor, id=id, company=company)
    contacts = vendor.contacts.all()

    if request.method == 'POST':
        form = VendorDetailsForm(request.POST, request.FILES, instance=vendor, company=company)
        if form.is_valid():
            form.save()

            vendor.contacts.all().delete()
            names = request.POST.getlist('contact_name[]')
            emails = request.POST.getlist('contact_email[]')
            phones = request.POST.getlist('contact_phone[]')

            for i in range(len(names)):
                if names[i]:
                    VendorContact.objects.create(
                        vendor=vendor,
                        name=names[i],
                        email=emails[i] if i < len(emails) else "",
                        phone=phones[i] if i < len(phones) else ""
                    )

            messages.success(request, "Vendor updated successfully.")
            return redirect('vendors_list')
    else:
        form = VendorDetailsForm(instance=vendor, company=company)

    return render(request, 'vendors/create.html', {
        'form': form,
        'company': company,
        'vendor': vendor,
        'contacts': contacts
    })


@login_required
def vendor_delete_view(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    vendor = get_object_or_404(Vendor, id=id, company=company)
    vendor.delete()
    messages.success(request, "Vendor deleted successfully.")
    return redirect('vendors_list')


@login_required
def vendor_import_view(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    if request.method == 'POST':
        import_file = request.FILES.get('import_file')
        if not import_file:
            messages.error(request, "Please upload a valid Excel file.")
            return redirect('vendor_import')

        import openpyxl
        try:
            wb = openpyxl.load_workbook(import_file, read_only=True)
            sheet = wb.active

            imported = 0
            skipped = 0
            failed = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue

                name = row[0]
                email = row[1] if len(row) > 1 else ""
                phone = row[2] if len(row) > 2 else ""
                currency = row[3] if len(row) > 3 else "INR"

                if not name:
                    failed += 1
                    continue

                if Vendor.objects.filter(company=company, vendor_name__iexact=name).exists():
                    skipped += 1
                    continue

                try:
                    Vendor.objects.create(
                        company=company,
                        vendor_name=name,
                        email=email,
                        phone=phone,
                        currency=currency,
                        created_by=request.user
                    )
                    imported += 1
                except Exception:
                    failed += 1

            messages.success(request, f"Import Summary - Imported: {imported}, Skipped: {skipped}, Failed: {failed}")
            return redirect('vendors_list')

        except Exception as e:
            messages.error(request, f"Failed to read Excel file: {str(e)}")
            return redirect('vendor_import')

    return render(request, 'vendors/import.html', {'company': company})


@login_required
def account_list_view(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    query = request.GET.get('q', '')
    sort_by = request.GET.get('sort', '-created_at')

    accounts = Account.objects.filter(company=company)
    if query:
        accounts = accounts.filter(
            name__icontains=query
        ) | accounts.filter(
            number__icontains=query
        ) | accounts.filter(
            bank_name__icontains=query
        ) | accounts.filter(
            bank_phone__icontains=query
        )

    if sort_by in ['name', '-name', 'opening_balance', '-opening_balance', 'created_at', '-created_at']:
        accounts = accounts.order_by(sort_by)

    paginator = Paginator(accounts, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate real-time current balances
    # pyrefly: ignore [missing-import]
    from django.db.models import Sum
    from account.models import IncomeTransaction, ExpenseTransaction, Transfer
    income_txs = IncomeTransaction.objects.filter(company=company)
    expense_txs = ExpenseTransaction.objects.filter(company=company)

    for acc in page_obj:
        inc_sum = income_txs.filter(account=acc).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        exp_sum = expense_txs.filter(account=acc).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        from_transfers = Transfer.objects.filter(company=company, from_account=acc).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        to_transfers = Transfer.objects.filter(company=company, to_account=acc).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        
        acc.current_balance = float(acc.opening_balance) + float(inc_sum) - float(exp_sum) - float(from_transfers) + float(to_transfers)

    return render(request, 'banking/accounts/list.html', {
        'company': company,
        'page_obj': page_obj,
        'query': query,
        'sort_by': sort_by
    })


@login_required
def account_create_view(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    if request.method == 'POST':
        form = AccountForm(request.POST, company=company)
        if form.is_valid():
            account = form.save(commit=False)
            account.company = company
            if account.default_account:
                Account.objects.filter(company=company, default_account=True).update(default_account=False)
            account.save()
            messages.success(request, "Account created successfully.")
            return redirect('account_list')
    else:
        form = AccountForm(company=company, initial={'default_account': False, 'type': 'Bank', 'opening_balance': '0.00'})

    return render(request, 'banking/accounts/form.html', {
        'form': form,
        'company': company
    })


@login_required
def account_edit_view(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    account = get_object_or_404(Account, id=id, company=company)

    if request.method == 'POST':
        form = AccountForm(request.POST, instance=account, company=company)
        if form.is_valid():
            acc = form.save(commit=False)
            if acc.default_account:
                Account.objects.filter(company=company, default_account=True).exclude(id=id).update(default_account=False)
            acc.save()
            messages.success(request, "Account updated successfully.")
            return redirect('account_list')
    else:
        form = AccountForm(instance=account, company=company)

    return render(request, 'banking/accounts/form.html', {
        'form': form,
        'company': company,
        'account': account
    })


@login_required
def account_delete_view(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    account = get_object_or_404(Account, id=id, company=company)
    account.delete()
    messages.success(request, "Account deleted successfully.")
    return redirect('account_list')


@login_required
def transaction_list(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
    from account.models import IncomeTransaction, ExpenseTransaction
    incomes = IncomeTransaction.objects.filter(company=company)
    expenses = ExpenseTransaction.objects.filter(company=company)
    
    transactions = []
    for inc in incomes:
        transactions.append({
            'id': inc.id,
            'type': 'Income',
            'date': inc.date,
            'number': inc.number,
            'account': inc.account.name if inc.account else '—',
            'contact': inc.customer.name if inc.customer else '—',
            'amount': inc.amount,
            'payment_method': inc.payment_method,
        })
    for exp in expenses:
        transactions.append({
            'id': exp.id,
            'type': 'Expense',
            'date': exp.date,
            'number': exp.number,
            'account': exp.account.name if exp.account else '—',
            'contact': exp.vendor.vendor_name if exp.vendor else '—',
            'amount': -exp.amount,
            'payment_method': exp.payment_method,
        })
    transactions.sort(key=lambda x: x['date'], reverse=True)
    records_exist = len(transactions) > 0
    return render(request, 'banking/transactions/transaction_list.html', {
        'company': company,
        'transactions': transactions,
        'records_exist': records_exist,
    })


@login_required
def new_income(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    last_income = IncomeTransaction.objects.filter(company=company).order_by('id').last()
    last_expense = ExpenseTransaction.objects.filter(company=company).order_by('id').last()
    
    last_num = 0
    for t in [last_income, last_expense]:
        if t and t.number.startswith('TRA-'):
            try:
                num = int(t.number.split('-')[1])
                if num > last_num:
                    last_num = num
            except Exception:
                pass
    next_num = last_num + 1
    next_tra_number = f"TRA-{next_num:05d}"

    form = IncomeTransactionForm(company=company, initial={
        'number': next_tra_number,
        'date': timezone.now().strftime('%Y-%m-%d'),
        'payment_method': 'Cash'
    })
    return render(request, 'banking/transactions/new_income.html', {
        'form': form,
        'company': company
    })


@login_required
def save_income(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    if request.method == 'POST':
        form = IncomeTransactionForm(request.POST, request.FILES, company=company)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.company = company
            transaction.save()
            messages.success(request, "Income Transaction created successfully.")
            return redirect('transaction_list')
        else:
            messages.error(request, "Please correct form errors.")
            return render(request, 'banking/transactions/new_income.html', {
                'form': form,
                'company': company
            })
    return redirect('new_income')


@login_required
def new_expense(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    last_income = IncomeTransaction.objects.filter(company=company).order_by('id').last()
    last_expense = ExpenseTransaction.objects.filter(company=company).order_by('id').last()
    
    last_num = 0
    for t in [last_income, last_expense]:
        if t and t.number.startswith('TRA-'):
            try:
                num = int(t.number.split('-')[1])
                if num > last_num:
                    last_num = num
            except Exception:
                pass
    next_num = last_num + 1
    next_tra_number = f"TRA-{next_num:05d}"

    form = ExpenseTransactionForm(company=company, initial={
        'number': next_tra_number,
        'date': timezone.now().strftime('%Y-%m-%d'),
        'payment_method': 'Cash'
    })
    return render(request, 'banking/transactions/new_expense.html', {
        'form': form,
        'company': company
    })


@login_required
def save_expense(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    if request.method == 'POST':
        form = ExpenseTransactionForm(request.POST, request.FILES, company=company)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.company = company
            transaction.save()
            messages.success(request, "Expense Transaction created successfully.")
            return redirect('transaction_list')
        else:
            messages.error(request, "Please correct form errors.")
            return render(request, 'banking/transactions/new_expense.html', {
                'form': form,
                'company': company
            })
    return redirect('new_expense')


@login_required
def transfers_landing(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    query = request.GET.get('q', '')
    sort_by = request.GET.get('sort', '-created_at')

    transfers = Transfer.objects.filter(company=company)
    if query:
        transfers = transfers.filter(
            transfer_number__icontains=query
        ) | transfers.filter(
            reference__icontains=query
        )

    if sort_by in ['transfer_number', '-transfer_number', 'amount', '-amount', 'date', '-date']:
        transfers = transfers.order_by(sort_by)

    paginator = Paginator(transfers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'banking/transfers/index.html', {
        'company': company,
        'page_obj': page_obj,
        'query': query,
        'sort_by': sort_by
    })


@login_required
def new_transfer(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    if request.method == 'POST':
        form = TransferForm(request.POST, request.FILES, company=company)
        if form.is_valid():
            transfer = form.save(commit=False)
            transfer.company = company
            transfer.created_by = request.user
            
            # Generate transfer number
            last_t = Transfer.objects.filter(company=company).order_by('id').last()
            next_num = 1
            if last_t and last_t.transfer_number.startswith('TRF-'):
                try:
                    next_num = int(last_t.transfer_number.split('-')[1]) + 1
                except Exception:
                    pass
            transfer.transfer_number = f"TRF-{next_num:05d}"
            
            transfer.save()
            messages.success(request, "Transfer created successfully.")
            return redirect('transfers_landing')
    else:
        # Default transfer number auto generation preview
        last_t = Transfer.objects.filter(company=company).order_by('id').last()
        next_num = 1
        if last_t and last_t.transfer_number.startswith('TRF-'):
            try:
                next_num = int(last_t.transfer_number.split('-')[1]) + 1
            except Exception:
                pass
        t_number = f"TRF-{next_num:05d}"
        
        form = TransferForm(company=company, initial={
            'date': timezone.now().strftime('%Y-%m-%d'),
            'payment_method': 'Cash'
        })

    return render(request, 'banking/transfers/new_transfer.html', {
        'form': form,
        'company': company
    })


@login_required
def edit_transfer(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    transfer = get_object_or_404(Transfer, id=id, company=company)

    if request.method == 'POST':
        form = TransferForm(request.POST, request.FILES, instance=transfer, company=company)
        if form.is_valid():
            form.save()
            messages.success(request, "Transfer updated successfully.")
            return redirect('transfers_landing')
    else:
        form = TransferForm(instance=transfer, company=company)

    return render(request, 'banking/transfers/new_transfer.html', {
        'form': form,
        'company': company,
        'transfer': transfer
    })


@login_required
def view_transfer(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    transfer = get_object_or_404(Transfer, id=id, company=company)
    return render(request, 'banking/transfers/view_transfer.html', {
        'company': company,
        'transfer': transfer
    })


@login_required
def delete_transfer(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    transfer = get_object_or_404(Transfer, id=id, company=company)
    transfer.delete()
    messages.success(request, "Transfer deleted successfully.")
    return redirect('transfers_landing')


import csv
# pyrefly: ignore [missing-import]
from django.http import HttpResponse

@login_required
def download_transfer_template(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="transfers_template.csv"'
    writer = csv.writer(response)
    writer.writerow(['From Account Number', 'To Account Number', 'Date (YYYY-MM-DD)', 'Amount', 'Payment Method', 'Reference', 'Description'])
    writer.writerow(['ACC-12345', 'ACC-54321', '2026-07-04', '1500.00', 'Bank', 'REF-001', 'Inter-account transfer'])
    return response


import openpyxl

@login_required
def import_transfers(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    if request.method == 'POST':
        import_file = request.FILES.get('import_file')
        if not import_file:
            messages.error(request, "Please upload a valid file.")
            return redirect('import_transfers')

        imported = 0
        skipped = 0
        errors = 0

        filename = import_file.name.lower()
        if filename.endswith('.csv'):
            try:
                decoded_file = import_file.read().decode('utf-8').splitlines()
                reader = csv.reader(decoded_file)
                next(reader, None) # skip header
                
                row_count = 0
                for row in reader:
                    row_count += 1
                    if row_count > 500:
                        messages.error(request, "Maximum of 500 rows allowed.")
                        return redirect('import_transfers')
                    
                    if not row or not any(row):
                        continue
                    
                    try:
                        from_acc_num = row[0]
                        to_acc_num = row[1]
                        date_str = row[2]
                        amount_val = row[3]
                        pay_method = row[4]
                        ref = row[5] if len(row) > 5 else ''
                        desc = row[6] if len(row) > 6 else ''
                        
                        from_acc = Account.objects.filter(company=company, number__iexact=from_acc_num).first()
                        to_acc = Account.objects.filter(company=company, number__iexact=to_acc_num).first()
                        
                        if not from_acc or not to_acc or from_acc == to_acc:
                            errors += 1
                            continue
                            
                        if ref and Transfer.objects.filter(company=company, reference__iexact=ref).exists():
                            skipped += 1
                            continue
                        
                        last_t = Transfer.objects.filter(company=company).order_by('id').last()
                        next_num = 1
                        if last_t and last_t.transfer_number.startswith('TRF-'):
                            try:
                                next_num = int(last_t.transfer_number.split('-')[1]) + 1
                            except Exception:
                                pass
                        t_number = f"TRF-{next_num:05d}"
                        
                        Transfer.objects.create(
                            company=company,
                            transfer_number=t_number,
                            from_account=from_acc,
                            to_account=to_acc,
                            date=date_str,
                            amount=amount_val, 
                            payment_method=pay_method,
                            reference=ref,
                            description=desc,
                            created_by=request.user
                        )
                        imported += 1
                    except Exception:
                        errors += 1
            except Exception as e:
                messages.error(request, f"Failed to process CSV file: {str(e)}")
                return redirect('import_transfers')
                
        elif filename.endswith(('.xls', '.xlsx')):
            try:
                wb = openpyxl.load_workbook(import_file, read_only=True)
                sheet = wb.active
                
                row_count = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_count += 1
                    if row_count > 500:
                        messages.error(request, "Maximum of 500 rows allowed.")
                        return redirect('import_transfers')
                    
                    if not row or not any(row):
                        continue
                        
                    try:
                        from_acc_num = str(row[0])
                        to_acc_num = str(row[1])
                        date_str = str(row[2])
                        amount_val = row[3]
                        pay_method = row[4]
                        ref = str(row[5]) if len(row) > 5 and row[5] else ''
                        desc = str(row[6]) if len(row) > 6 and row[6] else ''
                        
                        from_acc = Account.objects.filter(company=company, number__iexact=from_acc_num).first()
                        to_acc = Account.objects.filter(company=company, number__iexact=to_acc_num).first()
                        
                        if not from_acc or not to_acc or from_acc == to_acc:
                            errors += 1
                            continue
                            
                        if ref and Transfer.objects.filter(company=company, reference__iexact=ref).exists():
                            skipped += 1
                            continue
                            
                        last_t = Transfer.objects.filter(company=company).order_by('id').last()
                        next_num = 1
                        if last_t and last_t.transfer_number.startswith('TRF-'):
                            try:
                                next_num = int(last_t.transfer_number.split('-')[1]) + 1
                            except Exception:
                                pass
                        t_number = f"TRF-{next_num:05d}"
                        
                        Transfer.objects.create(
                            company=company,
                            transfer_number=t_number,
                            from_account=from_acc,
                            to_account=to_acc,
                            date=date_str,
                            amount=amount_val,
                            payment_method=pay_method,
                            reference=ref,
                            description=desc,
                            created_by=request.user
                        )
                        imported += 1
                    except Exception:
                        errors += 1
            except Exception as e:
                messages.error(request, f"Failed to process Excel file: {str(e)}")
                return redirect('import_transfers')
        else:
            messages.error(request, "Invalid file format. Please upload CSV, XLS or XLSX.")
            return redirect('import_transfers')

        messages.success(request, f"Import summary - Imported: {imported}, Skipped: {skipped}, Errors: {errors}")
        return redirect('transfers_landing')

    return render(request, 'banking/transfers/import_transfer.html', {'company': company})


@login_required
def reconciliation_landing(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    reconciliations = Reconciliation.objects.filter(account__company=company).order_by('-created_at')

    paginator = Paginator(reconciliations, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'banking/reconciliation/landing.html', {
        'company': company,
        'page_obj': page_obj
    })


@login_required
def new_reconciliation(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    form = ReconciliationForm(company=company)
    return render(request, 'banking/reconciliation/new.html', {
        'form': form,
        'company': company
    })


from django.http import HttpResponse

@login_required
def load_reconciliation_transactions(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return HttpResponse(status=400)

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    account_id = request.GET.get('account')

    if not start_date_str or not end_date_str or not account_id:
        return render(request, 'banking/reconciliation/partials/transaction_table.html', {'transactions': []})

    try:
        account = get_object_or_404(Account, id=account_id, company=company)
        
        incomes = IncomeTransaction.objects.filter(
            account=account,
            date__range=[start_date_str, end_date_str]
        )
        expenses = ExpenseTransaction.objects.filter(
            account=account,
            date__range=[start_date_str, end_date_str]
        )

        transactions_list = []
        for inc in incomes:
            transactions_list.append({
                'id': f"income_{inc.id}",
                'date': inc.date,
                'description': inc.description or 'Income Transaction',
                'contact': inc.customer.name if inc.customer else 'N/A',
                'deposit': inc.amount,
                'withdrawal': None,
            })
        for exp in expenses:
            transactions_list.append({
                'id': f"expense_{exp.id}",
                'date': exp.date,
                'description': exp.description or 'Expense Transaction',
                'contact': exp.vendor.name if exp.vendor else 'N/A',
                'deposit': None,
                'withdrawal': exp.amount,
            })

        transactions_list.sort(key=lambda x: x['date'])

        return render(request, 'banking/reconciliation/partials/transaction_table.html', {
            'transactions': transactions_list
        })
    except Exception as e:
        return HttpResponse(status=400)


@login_required
def save_reconciliation(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    if request.method == 'POST':
        form = ReconciliationForm(request.POST, company=company)
        if form.is_valid():
            reconciliation = form.save(commit=False)
            reconciliation.created_by = request.user
            reconciliation.save()

            cleared_ids = request.POST.getlist('cleared_transactions')
            
            for item_id in cleared_ids:
                if item_id.startswith('income_'):
                    inc_id = int(item_id.split('_')[1])
                    inc_tra = IncomeTransaction.objects.filter(id=inc_id, account__company=company).first()
                    if inc_tra:
                        ReconciliationTransaction.objects.create(
                            reconciliation=reconciliation,
                            income_transaction=inc_tra,
                            is_cleared=True
                        )
                elif item_id.startswith('expense_'):
                    exp_id = int(item_id.split('_')[1])
                    exp_tra = ExpenseTransaction.objects.filter(id=exp_id, account__company=company).first()
                    if exp_tra:
                        ReconciliationTransaction.objects.create(
                            reconciliation=reconciliation,
                            expense_transaction=exp_tra,
                            is_cleared=True
                        )
            
            messages.success(request, "Reconciliation saved successfully.")
            return redirect('reconciliation_landing')
        else:
            messages.error(request, "Please correct errors in form details.")
            return render(request, 'banking/reconciliation/new.html', {
                'form': form,
                'company': company
            })
    return redirect('new_reconciliation')


@login_required
def delete_reconciliation(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    reconciliation = get_object_or_404(Reconciliation, id=id, account__company=company)
    reconciliation.delete()
    messages.success(request, "Reconciliation deleted successfully.")
    return redirect('reconciliation_landing')


@login_required
def double_entry_landing(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')
    return render(request, 'double_entry/landing.html', {'company': company})


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@login_required
def reports_landing(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    if not Report.objects.exists():
        defaults = [
            ('Balance Sheet', 'Accounting', 'Snapshot of your business.', 'bi-layout-text-window-reverse'),
            ('General Ledger', 'Accounting', 'Detailed list of all transactions and their totals.', 'bi-journal-check'),
            ('Journal Entries', 'Accounting', 'Detailed list of all journal entries.', 'bi-journals'),
            ('Profit & Loss', 'Accounting', 'Quarterly profit and loss by category.', 'bi-graph-up-arrow'),
            ('Profit & Loss (COA)', 'Accounting', 'Quarterly profit and loss by chart of accounts.', 'bi-pie-chart'),
            ('Tax Summary', 'Accounting', 'Quarterly tax summary.', 'bi-percent'),
            ('Tax Summary (COA)', 'Accounting', 'Quarterly tax summary by chart of accounts.', 'bi-shield-check'),
            ('Trial Balance', 'Accounting', 'Balance of all chart of accounts.', 'bi-calculator'),
            ('Expense Summary', 'Income & Expense', 'Monthly expense summary by category.', 'bi-cart-dash'),
            ('Expense Summary (COA)', 'Income & Expense', 'Monthly expense summary by chart of accounts.', 'bi-cart'),
            ('Income Summary', 'Income & Expense', 'Monthly income summary by category.', 'bi-cash-coin'),
            ('Income Summary (COA)', 'Income & Expense', 'Monthly income summary by chart of accounts.', 'bi-currency-exchange'),
            ('Income vs Expense', 'Income & Expense', 'Monthly income vs expense by category.', 'bi-arrow-left-right'),
            ('Income vs Expense (COA)', 'Income & Expense', 'Monthly income vs expense by chart of accounts.', 'bi-columns-gap')
        ]
        for name, cat, desc, icon in defaults:
            Report.objects.create(name=name, category=cat, description=desc, icon=icon, is_system=True)

    pinned_relations = PinnedReport.objects.filter(user=request.user).order_by('display_order')
    pinned_reports = [pr.report for pr in pinned_relations]
    
    pinned_slots = []
    for i in range(6):
        if i < len(pinned_reports):
            pinned_slots.append(pinned_reports[i])
        else:
            pinned_slots.append(None)

    report_form = ReportForm()
    schedule_form = ScheduledReportForm()
    schedule_form.fields['report'].queryset = Report.objects.all()

    q = request.GET.get('q', '')
    cat_filter = request.GET.get('category', 'All')

    reports = Report.objects.all()
    if q:
        reports = reports.filter(
            name__icontains=q
        ) | reports.filter(
            category__icontains=q
        ) | reports.filter(
            description__icontains=q
        )

    if cat_filter != 'All':
        if cat_filter == 'Pinned':
            reports = reports.filter(pinned_by__user=request.user)
        elif cat_filter == 'Custom':
            reports = reports.filter(is_system=False)
        else:
            reports = reports.filter(category=cat_filter)

    accounting_reports = [r for r in reports if r.category == 'Accounting']
    income_expense_reports = [r for r in reports if r.category == 'Income & Expense']
    custom_reports = [r for r in reports if r.category == 'Custom' or not r.is_system]

    user_pinned_ids = set(PinnedReport.objects.filter(user=request.user).values_list('report_id', flat=True))

    return render(request, 'reports/index.html', {
        'company': company,
        'pinned_slots': pinned_slots,
        'accounting_reports': accounting_reports,
        'income_expense_reports': income_expense_reports,
        'custom_reports': custom_reports,
        'report_form': report_form,
        'schedule_form': schedule_form,
        'user_pinned_ids': user_pinned_ids,
        'q': q,
        'category': cat_filter
    })


@login_required
def create_report(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    if request.method == 'POST':
        form = ReportForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.is_system = False
            report.created_by = request.user
            report.save()
            messages.success(request, "Custom report created successfully.")
        else:
            messages.error(request, "Failed to create custom report.")
    return redirect('reports_landing')


@login_required
def update_report(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    report = get_object_or_404(Report, id=id, is_system=False)
    if request.method == 'POST':
        form = ReportForm(request.POST, instance=report)
        if form.is_valid():
            form.save()
            messages.success(request, "Report updated successfully.")
        else:
            messages.error(request, "Failed to update report.")
    return redirect('reports_landing')


@login_required
def delete_report(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    report = get_object_or_404(Report, id=id, is_system=False)
    report.delete()
    messages.success(request, "Custom report deleted successfully.")
    return redirect('reports_landing')


@login_required
def pin_report(request, id):
    report = get_object_or_404(Report, id=id)
    count = PinnedReport.objects.filter(user=request.user).count()
    if count >= 6:
        return JsonResponse({'status': 'error', 'message': 'Maximum 6 pinned reports allowed.'}, status=400)

    PinnedReport.objects.get_or_create(user=request.user, report=report, defaults={'display_order': count})
    return JsonResponse({'status': 'success'})


@login_required
def unpin_report(request, id):
    report = get_object_or_404(Report, id=id)
    PinnedReport.objects.filter(user=request.user, report=report).delete()
    return JsonResponse({'status': 'success'})


@login_required
def schedule_report(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    if request.method == 'POST':
        form = ScheduledReportForm(request.POST)
        if form.is_valid():
            schedule = form.save(commit=False)
            schedule.created_by = request.user
            schedule.save()
            messages.success(request, "Report scheduled successfully.")
        else:
            messages.error(request, "Failed to schedule report. Please check input values.")
    return redirect('reports_landing')


@login_required
def edit_report(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    report = get_object_or_404(Report, id=id)
    form = ReportEditForm(instance=report)
    return render(request, 'reports/report_edit.html', {
        'company': company,
        'report': report,
        'form': form
    })


@login_required
def update_report(request, id):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    report = get_object_or_404(Report, id=id)
    if request.method == 'POST':
        form = ReportEditForm(request.POST, instance=report)
        if form.is_valid():
            form.save()
            messages.success(request, "Report updated successfully.")
            return redirect('reports_landing')
        else:
            messages.error(request, "Please correct the errors in the form details.")
            return render(request, 'reports/report_edit.html', {
                'company': company,
                'report': report,
                'form': form
            })
    return redirect('reports_landing')


@login_required
def apps_marketplace(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    if not App.objects.exists():
        defaults = [
            ('Double Entry', 'double-entry', 'Professional double-entry accounting module.', 'bi-journal-text', '#565895', 'Accounting', 0.00, False, True, '1.2.0', 'Akaunting', 'Active'),
            ('Inventory Management', 'inventory-management', 'Manage products, stock, warehouses and inventory.', 'bi-box-seam', '#4f772d', 'Inventory', 6.00, True, True, '1.0.4', 'Akaunting', 'Active'),
            ('Reports & Analytics', 'reports-analytics', 'Generate accounting reports, dashboards and business insights.', 'bi-graph-up-arrow', '#e07a5f', 'Reports', 0.00, False, True, '2.1.0', 'Akaunting', 'Active'),
            ('Custom Fields', 'custom-fields', 'Add custom fields to your documents.', 'bi-file-earmark-plus', '#6366f1', 'Customization', 4.00, True, False, '1.0.0', 'Akaunting', 'Active'),
            ('Estimates', 'estimates', 'Send estimates/quotes to your customers.', 'bi-file-earmark-check', '#10b981', 'Sales', 5.00, True, False, '1.1.2', 'Akaunting', 'Active'),
            ('FluidPay', 'fluidpay', 'Automate payments using FluidPay.', 'bi-credit-card-2-front', '#0284c7', 'Payments', 4.00, True, False, '1.0.0', 'FluidPay Inc.', 'Active'),
            ('E-Invoice BE', 'e-invoice-be', 'Electronic invoicing integration for Belgium.', 'bi-file-earmark-text', '#eab308', 'Invoicing', 3.00, True, False, '1.0.0', 'Akaunting', 'Active'),
            ('Import Pro', 'import-pro', 'Import data using saved templates.', 'bi-cloud-arrow-up', '#ec4899', 'Import', 6.00, True, False, '2.0.0', 'Akaunting', 'Active')
        ]
        for name, slug, desc, logo, color, cat, price, is_paid, is_featured, ver, dev, status in defaults:
            App.objects.create(
                name=name, slug=slug, description=desc, logo=logo, banner_color=color,
                category=cat, price=price, is_paid=is_paid, is_featured=is_featured,
                version=ver, developer=dev, status=status
            )

    installed_app_ids = set(AppInstallation.objects.filter(user=request.user, installed=True).values_list('app_id', flat=True))

    q = request.GET.get('q', '')
    cat_filter = request.GET.get('category', 'All')

    apps = App.objects.all()
    if q:
        apps = apps.filter(name__icontains=q) | apps.filter(category__icontains=q) | apps.filter(developer__icontains=q)

    if cat_filter != 'All':
        if cat_filter == 'Paid':
            apps = apps.filter(is_paid=True)
        elif cat_filter == 'Free':
            apps = apps.filter(is_paid=False)
        elif cat_filter == 'Installed':
            apps = apps.filter(id__in=installed_app_ids)
        elif cat_filter == 'Featured':
            apps = apps.filter(is_featured=True)

    top_paid = [a for a in apps if a.is_paid][:4]
    new_apps = [a for a in apps if not a.is_paid][:4]
    top_free = [a for a in apps if not a.is_paid][:4]

    return render(request, 'apps/marketplace.html', {
        'company': company,
        'top_paid': top_paid,
        'new_apps': new_apps,
        'top_free': top_free,
        'installed_app_ids': installed_app_ids,
        'q': q,
        'category': cat_filter
    })


@login_required
def app_detail(request, slug):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    app = get_object_or_404(App, slug=slug)
    installed = AppInstallation.objects.filter(user=request.user, app=app, installed=True).exists()

    return render(request, 'apps/app_detail.html', {
        'company': company,
        'app': app,
        'installed': installed
    })


@login_required
def install_app(request, id):
    app = get_object_or_404(App, id=id)
    inst, created = AppInstallation.objects.get_or_create(user=request.user, app=app)
    inst.installed = True
    inst.save()
    messages.success(request, f"App {app.name} installed successfully.")
    return redirect('apps_marketplace')


@login_required
def open_app(request, id):
    app = get_object_or_404(App, id=id)
    installed = AppInstallation.objects.filter(user=request.user, app=app, installed=True).exists()
    if not installed:
        messages.error(request, "Please install the app first.")
        return redirect('apps_marketplace')

    if 'Double Entry' in app.name or app.slug == 'double-entry':
        return redirect('double_entry_landing')
    elif 'Reports' in app.name or app.slug == 'reports-analytics':
        return redirect('reports_landing')
    elif 'Inventory' in app.name or app.slug == 'inventory-management':
        return redirect('item_list')
    else:
        messages.success(request, f"App {app.name} opened.")
        return redirect('apps_marketplace')


@login_required
def widget_detail_api(request, id):
    widget = get_object_or_404(DashboardWidget, id=id, user=request.user)
    return JsonResponse({
        'id': widget.id,
        'widget_name': widget.widget_name,
        'widget_type': widget.widget_type,
        'width': widget.width,
        'sort_order': widget.sort_order,
        'is_active': widget.is_active,
    })


@login_required
def widget_update_api(request, id):
    widget = get_object_or_404(DashboardWidget, id=id, user=request.user)
    if request.method == 'POST':
        form = DashboardWidgetForm(request.POST, instance=widget, user=request.user)
        if form.is_valid():
            form.save()
            return JsonResponse({
                'status': 'success',
                'message': 'Widget updated successfully.',
                'widget': {
                    'id': widget.id,
                    'widget_name': widget.widget_name,
                    'widget_type': widget.widget_type,
                    'width': widget.width,
                    'sort_order': widget.sort_order,
                }
            })
        else:
            return JsonResponse({
                'status': 'error',
                'errors': form.errors.get_json_data()
            }, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=400)


@login_required
def new_dashboard(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    form = DashboardForm(user=request.user)
    return render(request, 'dashboard/new_dashboard.html', {
        'company': company,
        'form': form
    })


@login_required
def save_dashboard(request):
    try:
        company = request.user.company
    except Company.DoesNotExist:
        return redirect('company_setup')

    if request.method == 'POST':
        form = DashboardForm(request.POST, user=request.user)
        if form.is_valid():
            db = form.save(commit=False)
            db.created_by = request.user
            db.save()
            form.save_m2m()
            messages.success(request, f"Dashboard '{db.name}' created successfully.")
            return redirect('dashboard')
        else:
            messages.error(request, "Failed to create dashboard. Please correct the errors.")
            return render(request, 'dashboard/new_dashboard.html', {
                'company': company,
                'form': form
            })
    return redirect('new_dashboard')


@login_required
def mark_notification_read(request, id):
    notif = get_object_or_404(Notification, id=id, recipient=request.user)
    notif.read_status = True
    notif.save()
    return JsonResponse({
        'status': 'success',
        'unread_count': Notification.objects.filter(recipient=request.user, read_status=False).count()
    })











